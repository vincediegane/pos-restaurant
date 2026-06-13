import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
IMPORTS_DIR = ROOT / "imports"
PAYLOAD_DIR = ROOT / "addons" / "restaurant_starter" / "import_payload"
IMAGES_DIR = PAYLOAD_DIR / "images"
EXCEL_PATH = IMPORTS_DIR / "restaurant_products_100.xlsx"
JSON_PATH = PAYLOAD_DIR / "products.json"


CATEGORIES = {
    "Plats": ["Riz", "Poulet", "Poisson", "Yassa", "Mafe", "Burger", "Pizza", "Pasta", "Tacos", "Grillade"],
    "Boissons": ["Eau", "Jus", "Soda", "Cafe", "The", "Bissap", "Gingembre", "Lait", "Smoothie", "Citronnade"],
    "Desserts": ["Gateau", "Crepe", "Glace", "Tarte", "Mousse", "Fruit", "Yaourt", "Beignet", "Pancake", "Brownie"],
    "Entrees": ["Salade", "Soupe", "Nems", "Bruschetta", "Tapas", "Ailes", "Brochette", "Pastel", "Accra", "Samoussa"],
    "Ingredients": ["Riz sac", "Huile", "Oignon", "Tomate", "Farine", "Sucre", "Sel", "Epice", "Laitue", "Pomme terre"],
}

PALETTES = [
    ("#f7efe4", "#bc4749", "#386641"),
    ("#edf6f9", "#006d77", "#e29578"),
    ("#fff3b0", "#335c67", "#9e2a2b"),
    ("#f1faee", "#457b9d", "#e63946"),
    ("#fefae0", "#606c38", "#dda15e"),
]


def get_font(size):
    for font_name in ["arial.ttf", "DejaVuSans-Bold.ttf"]:
        try:
            return ImageFont.truetype(font_name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def draw_product_image(product, image_path, palette):
    bg, main, accent = palette
    img = Image.new("RGB", (420, 320), bg)
    draw = ImageDraw.Draw(img)

    draw.rounded_rectangle((24, 24, 396, 296), radius=24, fill=bg, outline=main, width=6)
    draw.ellipse((115, 70, 305, 235), fill=main)
    draw.ellipse((148, 98, 272, 205), fill=accent)
    draw.rectangle((174, 205, 246, 246), fill=accent)

    title = product["name"][:24]
    sku = product["default_code"]
    title_font = get_font(30)
    sku_font = get_font(18)
    title_box = draw.textbbox((0, 0), title, font=title_font)
    sku_box = draw.textbbox((0, 0), sku, font=sku_font)
    draw.text(((420 - (title_box[2] - title_box[0])) / 2, 250), title, fill="#222222", font=title_font)
    draw.text(((420 - (sku_box[2] - sku_box[0])) / 2, 36), sku, fill="#222222", font=sku_font)

    img.save(image_path, "PNG")


def build_products():
    products = []
    index = 1
    for category, names in CATEGORIES.items():
        for name in names:
            for variant in ["Classique", "Premium"]:
                is_stock_item = category == "Ingredients"
                price = 0 if is_stock_item else 500 + (index * 125)
                cost = 300 + (index * 65)
                product_name = f"{name} {variant}"
                products.append(
                    {
                        "name": product_name,
                        "default_code": f"RESTO-{index:03d}",
                        "barcode": f"230000000{index:03d}",
                        "category": f"Restaurant / {category}",
                        "pos_category": category if not is_stock_item else "",
                        "sale_price": price,
                        "cost": cost,
                        "detailed_type": "product",
                        "available_in_pos": not is_stock_item,
                        "can_be_sold": not is_stock_item,
                        "can_be_purchased": True,
                        "image_file": f"images/RESTO-{index:03d}.png",
                    }
                )
                index += 1
    return products


def write_excel(products):
    wb = Workbook()
    ws = wb.active
    ws.title = "Produits Restaurant"
    headers = [
        "Nom",
        "Reference interne",
        "Code-barres",
        "Categorie produit",
        "Categorie POS",
        "Prix de vente",
        "Cout",
        "Type",
        "Disponible POS",
        "Peut etre vendu",
        "Peut etre achete",
        "Fichier image",
        "Image",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F4858")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, product in enumerate(products, start=2):
        ws.append(
            [
                product["name"],
                product["default_code"],
                product["barcode"],
                product["category"],
                product["pos_category"],
                product["sale_price"],
                product["cost"],
                product["detailed_type"],
                product["available_in_pos"],
                product["can_be_sold"],
                product["can_be_purchased"],
                product["image_file"],
                "",
            ]
        )
        img = ExcelImage(str(PAYLOAD_DIR / product["image_file"]))
        img.width = 92
        img.height = 70
        ws.add_image(img, f"M{row_idx}")
        ws.row_dimensions[row_idx].height = 58

    widths = [28, 18, 18, 24, 18, 14, 12, 12, 16, 16, 16, 24, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(products) + 1}"
    wb.save(EXCEL_PATH)


def main():
    IMPORTS_DIR.mkdir(exist_ok=True)
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    products = build_products()
    for i, product in enumerate(products):
        image_path = PAYLOAD_DIR / product["image_file"]
        draw_product_image(product, image_path, PALETTES[i % len(PALETTES)])

    JSON_PATH.write_text(json.dumps(products, ensure_ascii=True, indent=2), encoding="utf-8")
    write_excel(products)
    print(EXCEL_PATH)
    print(JSON_PATH)
    print(len(products))


if __name__ == "__main__":
    main()
